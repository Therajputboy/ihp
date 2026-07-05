# Admin Route Report Module

from flask import Blueprint, request, jsonify, make_response
from utils.exceptionlogging import ExceptionLogging
import traceback
from utils.jwt import jwt_required
from datetime import datetime, timedelta
import pytz
from utils import db
from utils.schemas import users, routes as route_table, driver_routes, log_table
from utils.globalconstants import CustomException
from utils.logging import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from io import BytesIO
from flask import send_file

bp_route_report = Blueprint('bp_route_report', __name__)

@bp_route_report.route('/route-report/summary', methods=['GET'])
@jwt_required
def get_route_report_summary():
    """
    Get route report summary statistics only
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access route reports.")

        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')

        # Parse dates
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date_str else None
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date_str else None

        if to_date:
            to_date = to_date.replace(hour=23, minute=59, second=59)

        # Get all routes
        all_routes = db.get_all(route_table.table_name, route_table.json_fields)

        # Get all driver routes for trip information
        all_driver_routes = db.get_by_filter(
            driver_routes.table_name,
            [["status", "IN", ["completed", "active", "scheduled"]]],
            driver_routes.json_fields,
            order=["-created_at"]
        )

        # Build route map with trip counts
        route_latest_status_set = set()
        route_trips_map = {}
        for driver_route in all_driver_routes:
            route_id_key = driver_route.get('route_id', '')

            # Filter by date range based on driver route's updated_at
            updated_at = driver_route.get('updated_at')
            if updated_at:
                if from_date and updated_at < from_date:
                    continue
                if to_date and updated_at > to_date:
                    continue

            if route_id_key not in route_trips_map:
                route_trips_map[route_id_key] = {
                    'total_trips': 0,
                    'completed_trips': 0,
                    'in_progress_trips': 0,
                    'latest_status': ''
                }

            status = driver_route.get('status', '')
            route_trips_map[route_id_key]['total_trips'] += 1

            # Only set latest_status for the first (most recent) entry of each route
            if route_id_key not in route_latest_status_set:
                route_trips_map[route_id_key]['latest_status'] = status
                route_latest_status_set.add(route_id_key)

            if status == 'completed':
                route_trips_map[route_id_key]['completed_trips'] += 1
            elif status == 'active':
                route_trips_map[route_id_key]['in_progress_trips'] += 1

        # Build summary statistics
        total_distance = 0.0
        total_trips = 0
        route_ids_set = set()

        for route in all_routes:
            route_id_key = route.get('route_id', '')

            # Filter by date range
            created_at = route.get('created_at')
            if created_at:
                if from_date and created_at < from_date:
                    continue
                if to_date and created_at > to_date:
                    continue

            route_ids_set.add(route_id_key)

            # Get trip information
            trips_info = route_trips_map.get(route_id_key, {
                'total_trips': 0,
                'completed_trips': 0,
                'in_progress_trips': 0
            })

            total_trips_count = trips_info.get('total_trips', 0)

            # Get distance information
            distance = route.get('distance', 0)
            total_distance += distance if distance else 0

            total_trips += total_trips_count

        payload.update({
            "message": "Route report summary fetched successfully.",
            "summary": {
                "total_routes": len(route_ids_set),
                "total_trips": total_trips,
                "total_distance_km": round(total_distance, 2)
            },
            "from_date": from_date_str,
            "to_date": to_date_str
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

@bp_route_report.route('/route-report', methods=['GET'])
@jwt_required
def get_route_report():
    """
    Get combined route report with summary statistics and detailed report data
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - route_id: specific route ID (optional, if not provided fetches all routes)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access route reports.")

        # Get filters
        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')
        route_id = request.args.get('route_id', '')

        # Parse dates
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date_str else None
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date_str else None

        if to_date:
            to_date = to_date.replace(hour=23, minute=59, second=59)

        # Get all routes
        all_routes = db.get_all(route_table.table_name, route_table.json_fields)

        # Get all driver routes for trip information (newest first to get latest status)
        all_driver_routes = db.get_by_filter(
            driver_routes.table_name,
            [["status", "IN", ["completed", "active", "scheduled"]]],
            driver_routes.json_fields,
            order=["-created_at"]
        )

        # Build route map with trip counts (with date filtering)
        # Track which routes we've already set the latest status for
        route_latest_status_set = set()
        route_trips_map = {}
        for driver_route in all_driver_routes:
            route_id_key = driver_route.get('route_id', '')

            # Filter by date range based on driver route's updated_at
            updated_at = driver_route.get('updated_at')
            if updated_at:
                if from_date and updated_at < from_date:
                    continue
                if to_date and updated_at > to_date:
                    continue

            if route_id_key not in route_trips_map:
                route_trips_map[route_id_key] = {
                    'total_trips': 0,
                    'completed_trips': 0,
                    'in_progress_trips': 0,
                    'latest_status': ''
                }

            status = driver_route.get('status', '')
            route_trips_map[route_id_key]['total_trips'] += 1

            # Only set latest_status for the first (most recent) entry of each route
            if route_id_key not in route_latest_status_set:
                route_trips_map[route_id_key]['latest_status'] = status
                route_latest_status_set.add(route_id_key)

            if status == 'completed':
                route_trips_map[route_id_key]['completed_trips'] += 1
            elif status == 'active':
                route_trips_map[route_id_key]['in_progress_trips'] += 1

        # Build report data and calculate statistics
        report_data = []
        total_distance = 0.0
        total_trips = 0
        route_ids_set = set()

        for route in all_routes:
            route_id_key = route.get('route_id', '')

            # Filter by specific route if provided
            if route_id and route_id_key != route_id:
                continue

            # Filter by date range
            created_at = route.get('created_at')
            if created_at:
                if from_date and created_at < from_date:
                    continue
                if to_date and created_at > to_date:
                    continue

            route_ids_set.add(route_id_key)

            # Get trip information
            trips_info = route_trips_map.get(route_id_key, {
                'total_trips': 0,
                'completed_trips': 0,
                'in_progress_trips': 0
            })

            total_trips_count = trips_info.get('total_trips', 0)
            completed_trips_count = trips_info.get('completed_trips', 0)
            in_progress_trips_count = trips_info.get('in_progress_trips', 0)

            # Get distance information
            distance = route.get('distance', 0)
            total_distance += distance if distance else 0

            # Get route status from latest driver route
            route_status = trips_info.get('latest_status', 'Active').capitalize()

            total_trips += total_trips_count

            report_data.append({
                'route_name': route.get('route_name', 'N/A'),
                'total_distance_km': round(float(distance) if distance else 0, 2),
                'total_trips': total_trips_count,
                'completed_trips': completed_trips_count,
                'in_progress_trips': in_progress_trips_count,
                'status': route_status,
                'route_id': route_id_key,
                'state': route.get('state', ''),
                'city': route.get('city', '')
            })

        payload.update({
            "message": "Route report fetched successfully.",
            "summary": {
                "total_routes": len(route_ids_set),
                "total_trips": total_trips,
                "total_distance_km": round(total_distance, 2)
            },
            "report_data": report_data,
            "total_records": len(report_data),
            "from_date": from_date_str,
            "to_date": to_date_str,
            "route_id_filter": route_id
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

@bp_route_report.route('/route-report/export/excel', methods=['GET'])
@jwt_required
def export_route_report_excel():
    """
    Export route report to Excel format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - route_id: specific route ID (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export route reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        route_id = request.args.get('route_id', '')

        # Build report data
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date else None
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date else None
        if to_date_obj:
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)

        all_routes = db.get_all(route_table.table_name, route_table.json_fields)
        all_driver_routes = db.get_by_filter(
            driver_routes.table_name,
            [["status", "IN", ["completed", "active", "scheduled"]]],
            driver_routes.json_fields,
            order=["-created_at"]
        )

        route_latest_status_set = set()
        route_trips_map = {}
        for driver_route in all_driver_routes:
            route_id_key = driver_route.get('route_id', '')

            # Filter by date range based on driver route's updated_at
            updated_at = driver_route.get('updated_at')
            if updated_at:
                if from_date_obj and updated_at < from_date_obj:
                    continue
                if to_date_obj and updated_at > to_date_obj:
                    continue

            if route_id_key not in route_trips_map:
                route_trips_map[route_id_key] = {
                    'total_trips': 0,
                    'completed_trips': 0,
                    'in_progress_trips': 0,
                    'latest_status': ''
                }

            status = driver_route.get('status', '')
            route_trips_map[route_id_key]['total_trips'] += 1

            # Only set latest_status for the first (most recent) entry of each route
            if route_id_key not in route_latest_status_set:
                route_trips_map[route_id_key]['latest_status'] = status
                route_latest_status_set.add(route_id_key)

            if status == 'completed':
                route_trips_map[route_id_key]['completed_trips'] += 1
            elif status == 'active':
                route_trips_map[route_id_key]['in_progress_trips'] += 1

        report_data = []
        for route in all_routes:
            route_id_key = route.get('route_id', '')

            if route_id and route_id_key != route_id:
                continue

            created_at = route.get('created_at')
            if created_at:
                if from_date_obj and created_at < from_date_obj:
                    continue
                if to_date_obj and created_at > to_date_obj:
                    continue

            trips_info = route_trips_map.get(route_id_key, {
                'total_trips': 0,
                'completed_trips': 0,
                'in_progress_trips': 0
            })

            distance = route.get('distance', 0)
            route_status = trips_info.get('latest_status', 'Active').capitalize()

            report_data.append({
                'route_name': route.get('route_name', 'N/A'),
                'total_distance_km': round(float(distance) if distance else 0, 2),
                'total_trips': trips_info.get('total_trips', 0),
                'completed_trips': trips_info.get('completed_trips', 0),
                'in_progress_trips': trips_info.get('in_progress_trips', 0),
                'status': route_status,
                'route_id': route_id_key
            })

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Route Report"

        # Add header with report info
        ws['A1'] = "Route Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:I1')

        # Add date range
        ws['A2'] = f"Report Date Range: {from_date} to {to_date}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:I2')

        # Add summary statistics
        total_routes = len(set([item.get('route_id', '') for item in report_data if item.get('route_id')]))
        total_trips = sum([item.get('total_trips', 0) for item in report_data])
        total_distance = sum([item.get('total_distance_km', 0) for item in report_data])

        ws['A3'] = "Summary Statistics"
        ws['A3'].font = Font(size=12, bold=True)

        ws['A4'] = "Total Routes:"
        ws['B4'] = total_routes
        ws['A5'] = "Total Trips:"
        ws['B5'] = total_trips
        ws['A6'] = "Total Distance (KM):"
        ws['B6'] = round(total_distance, 2)

        # Add table header
        headers = ['Route Name', 'Total Distance (KM)', 'Total Trips', 'Completed Trips',
                   'In Progress Trips', 'Status']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add data rows
        for row_num, item in enumerate(report_data, 9):
            ws.cell(row=row_num, column=1).value = item.get('route_name', 'N/A')
            ws.cell(row=row_num, column=2).value = item.get('total_distance_km', 0)
            ws.cell(row=row_num, column=3).value = item.get('total_trips', 0)
            ws.cell(row=row_num, column=4).value = item.get('completed_trips', 0)
            ws.cell(row=row_num, column=5).value = item.get('in_progress_trips', 0)
            ws.cell(row=row_num, column=6).value = item.get('status', 'Active')

            for col_num in range(1, 7):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'route_report_{from_date}_to_{to_date}.xlsx'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

@bp_route_report.route('/route-report/export/pdf', methods=['GET'])
@jwt_required
def export_route_report_pdf():
    """
    Export route report to PDF format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - route_id: specific route ID (optional)
    - company_logo: URL to company logo (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export route reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        route_id = request.args.get('route_id', '')
        company_logo = request.args.get('company_logo', '')

        # Build report data
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date else None
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date else None
        if to_date_obj:
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)

        all_routes = db.get_all(route_table.table_name, route_table.json_fields)
        all_driver_routes = db.get_by_filter(
            driver_routes.table_name,
            [["status", "IN", ["completed", "active", "scheduled"]]],
            driver_routes.json_fields,
            order=["-created_at"]
        )

        route_latest_status_set = set()
        route_trips_map = {}
        for driver_route in all_driver_routes:
            route_id_key = driver_route.get('route_id', '')

            # Filter by date range based on driver route's updated_at
            updated_at = driver_route.get('updated_at')
            if updated_at:
                if from_date_obj and updated_at < from_date_obj:
                    continue
                if to_date_obj and updated_at > to_date_obj:
                    continue

            if route_id_key not in route_trips_map:
                route_trips_map[route_id_key] = {
                    'total_trips': 0,
                    'completed_trips': 0,
                    'in_progress_trips': 0,
                    'latest_status': ''
                }

            status = driver_route.get('status', '')
            route_trips_map[route_id_key]['total_trips'] += 1

            # Only set latest_status for the first (most recent) entry of each route
            if route_id_key not in route_latest_status_set:
                route_trips_map[route_id_key]['latest_status'] = status
                route_latest_status_set.add(route_id_key)

            if status == 'completed':
                route_trips_map[route_id_key]['completed_trips'] += 1
            elif status == 'active':
                route_trips_map[route_id_key]['in_progress_trips'] += 1

        report_data = []
        for route in all_routes:
            route_id_key = route.get('route_id', '')

            if route_id and route_id_key != route_id:
                continue

            created_at = route.get('created_at')
            if created_at:
                if from_date_obj and created_at < from_date_obj:
                    continue
                if to_date_obj and created_at > to_date_obj:
                    continue

            trips_info = route_trips_map.get(route_id_key, {
                'total_trips': 0,
                'completed_trips': 0,
                'in_progress_trips': 0
            })

            distance = route.get('distance', 0)
            route_status = trips_info.get('latest_status', 'Active').capitalize()

            report_data.append({
                'route_name': route.get('route_name', 'N/A'),
                'total_distance_km': round(float(distance) if distance else 0, 2),
                'total_trips': trips_info.get('total_trips', 0),
                'completed_trips': trips_info.get('completed_trips', 0),
                'in_progress_trips': trips_info.get('in_progress_trips', 0),
                'status': route_status,
                'route_id': route_id_key
            })

        # Create PDF
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch, pageCompression=1)
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=10,
            alignment=1
        )

        # Add company logo if provided
        if company_logo:
            try:
                logo = Image(company_logo, width=1*inch, height=1*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass

        # Add title
        elements.append(Paragraph("Route Report", title_style))
        elements.append(Paragraph(f"<b>Report Date Range:</b> {from_date} to {to_date}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Add summary statistics
        total_routes = len(set([item.get('route_id', '') for item in report_data if item.get('route_id')]))
        total_trips = sum([item.get('total_trips', 0) for item in report_data])
        total_distance = sum([item.get('total_distance_km', 0) for item in report_data])

        summary_text = f"""
        <b>Summary Statistics</b><br/>
        Total Routes: {total_routes}<br/>
        Total Trips: {total_trips}<br/>
        Total Distance: {round(total_distance, 2)} KM
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Create report table with wrapped text
        normal_style = styles['Normal']
        normal_style.fontSize = 8
        normal_style.leading = 10

        table_data = [[Paragraph(text, normal_style) for text in ['Route Name', 'Total Distance (KM)', 'Total Trips', 'Completed Trips',
                      'In Progress Trips', 'Status']]]

        for item in report_data:
            table_data.append([
                Paragraph(str(item.get('route_name', 'N/A')), normal_style),
                Paragraph(f"{item.get('total_distance_km', 0)} KM", normal_style),
                Paragraph(str(item.get('total_trips', 0)), normal_style),
                Paragraph(str(item.get('completed_trips', 0)), normal_style),
                Paragraph(str(item.get('in_progress_trips', 0)), normal_style),
                Paragraph(str(item.get('status', 'Active')), normal_style)
            ])

        # Create table
        table = Table(table_data, colWidths=[1.8*inch, 1.3*inch, 1.0*inch, 1.3*inch,
                                             1.4*inch, 1.0*inch])

        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)

        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'route_report_{from_date}_to_{to_date}.pdf'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)
