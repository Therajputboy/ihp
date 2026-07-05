# Admin Truck Report Module

from flask import Blueprint, request, jsonify, make_response
from utils.exceptionlogging import ExceptionLogging
import traceback
from utils.jwt import jwt_required
from datetime import datetime, timedelta
import pytz
from utils import db
from utils.schemas import users, routes as route_table, driver_routes, trucks, log_table
from utils.globalconstants import CustomException
from utils.logging import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from io import BytesIO
from flask import send_file

bp_truck_report = Blueprint('bp_truck_report', __name__)

@bp_truck_report.route('/truck-report/summary', methods=['GET'])
@jwt_required
def get_truck_report_summary():
    """
    Get truck report summary statistics only
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access truck reports.")

        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')

        # Parse dates
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date_str else None
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date_str else None

        if to_date:
            to_date = to_date.replace(hour=23, minute=59, second=59)

        # Get all trucks
        all_trucks = db.get_all(trucks.table_name, trucks.json_fields)
        trucks_map = {t.get('truckid', ''): t for t in all_trucks}

        # Get all driver routes
        all_driver_routes = db.get_all(driver_routes.table_name, driver_routes.json_fields)

        # Get completed routes
        completed_routes = db.get_by_filter(
            route_table.table_name,
            [["status", "=", "completed"], ["approved", "=", 1]],
            route_table.json_fields,
            order=["-created_at"]
        )

        route_map = {route.get('route_id', ''): route for route in completed_routes}

        total_distance = 0.0
        truck_ids_set = set()

        for driver_route in all_driver_routes:
            truckid = driver_route.get('truckid', '')

            if not truckid:
                continue

            truck_ids_set.add(truckid)

            route_id = driver_route.get('route_id', '')
            route = route_map.get(route_id, {})

            if not route:
                continue

            # Filter by date range
            completed_at = route.get('completed_at')
            if completed_at:
                if from_date and completed_at < from_date:
                    continue
                if to_date and completed_at > to_date:
                    continue

            truck = trucks_map.get(truckid, {})
            actual_distance = truck.get('actual_distance', truck.get('distance', 0))
            total_distance += actual_distance if actual_distance else 0

        payload.update({
            "message": "Truck report summary fetched successfully.",
            "summary": {
                "total_trucks": len(truck_ids_set),
                "total_distance_km": round(total_distance, 2)
            },
            "from_date": from_date_str,
            "to_date": to_date_str
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

def get_driver_name(driverid, users_map):
    """Helper to get driver name from users map"""
    user = users_map.get(driverid, {})
    return user.get('name', 'N/A')

def calculate_distance_difference(total_distance, actual_distance):
    """Calculate extra distance travelled"""
    try:
        if not total_distance or not actual_distance:
            return 0.0
        total = float(total_distance) if isinstance(total_distance, str) else total_distance
        actual = float(actual_distance) if isinstance(actual_distance, str) else actual_distance
        return round(actual - total, 2)
    except:
        return 0.0

@bp_truck_report.route('/truck-report', methods=['GET'])
@jwt_required
def get_truck_report():
    """
    Get combined truck report with summary statistics and detailed report data
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - truck_id: specific truck ID (optional, if not provided fetches all trucks)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access truck reports.")

        # Get filters
        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')
        truck_id = request.args.get('truck_id', '')

        # Parse dates
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date_str else None
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date_str else None

        if to_date:
            to_date = to_date.replace(hour=23, minute=59, second=59)

        # Get all users for mapping
        all_users = db.get_all(users.table_name, users.json_fields)
        users_map = {u.get('userid', ''): u for u in all_users}

        # Get all trucks
        all_trucks = db.get_all(trucks.table_name, trucks.json_fields)
        trucks_map = {t.get('truckid', ''): t for t in all_trucks}

        # Get completed driver routes (these have truck assignments)
        all_driver_routes = db.get_by_filter(
            driver_routes.table_name,
            [["status", "=", "completed"]],
            driver_routes.json_fields,
            order=["-updated_at"]
        )

        # Get completed routes for additional details
        completed_routes = db.get_by_filter(
            route_table.table_name,
            [["status", "=", "completed"], ["approved", "=", 1]],
            route_table.json_fields,
            order=["-created_at"]
        )

        # Build route map
        route_map = {route.get('route_id', ''): route for route in completed_routes}

        # Build report data and calculate statistics
        report_data = []
        total_distance = 0.0
        truck_ids_set = set()

        for driver_route in all_driver_routes:
            truckid = driver_route.get('truckid', '')

            if not truckid:
                continue

            # Filter by specific truck if provided
            if truck_id and truckid != truck_id:
                continue

            truck_ids_set.add(truckid)

            route_id = driver_route.get('route_id', '')
            route = route_map.get(route_id, {})

            if not route:
                continue

            # Filter by date range
            completed_at = route.get('completed_at')
            if completed_at:
                if from_date and completed_at < from_date:
                    continue
                if to_date and completed_at > to_date:
                    continue

            truck = trucks_map.get(truckid, {})
            truck_number = truck.get('registration_number') or truck.get('truck_number') or truckid or 'N/A'

            # Get driver information
            driverids = driver_route.get('driverid', [])
            if not isinstance(driverids, list):
                driverids = [driverids] if driverids else []

            driver1 = get_driver_name(driverids[0], users_map) if len(driverids) > 0 else 'N/A'
            driver2 = get_driver_name(driverids[1], users_map) if len(driverids) > 1 else 'N/A'

            # Get distance information
            route_distance = truck.get('distance', 0)
            actual_distance = truck.get('actual_distance', route_distance)
            extra_distance = calculate_distance_difference(route_distance, actual_distance)

            # Update statistics
            total_distance += actual_distance if actual_distance else 0

            # Get damage and status information
            damage_reported = truck.get('damage_reported', False)
            damage_status = "Yes" if damage_reported else "No"
            truck_status = truck.get('status', 'Active')

            report_data.append({
                'truck_number': truck_number,
                'date': completed_at.strftime('%d/%m/%Y') if completed_at else 'N/A',
                'route_assigned': route.get('route_name', 'N/A'),
                'driver_1': driver1,
                'driver_2': driver2,
                'total_distance_km': round(float(route_distance) if route_distance else 0, 2),
                'actual_distance_km': round(float(actual_distance) if actual_distance else 0, 2),
                'extra_distance_km': extra_distance,
                'damage_reported': damage_status,
                'status': truck_status,
                'truckid': truckid,
                'route_id': route_id
            })

        payload.update({
            "message": "Truck report fetched successfully.",
            "summary": {
                "total_trucks": len(truck_ids_set),
                "total_trips": len(report_data),
                "total_distance_km": round(total_distance, 2)
            },
            "report_data": report_data,
            "total_records": len(report_data),
            "from_date": from_date_str,
            "to_date": to_date_str,
            "truck_id_filter": truck_id
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

@bp_truck_report.route('/truck-report/export/excel', methods=['GET'])
@jwt_required
def export_truck_report_excel():
    """
    Export truck report to Excel format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - truck_id: specific truck ID (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export truck reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        truck_id = request.args.get('truck_id', '')

        # Build report data
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date else None
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date else None
        if to_date_obj:
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)

        all_users = db.get_all(users.table_name, users.json_fields)
        users_map = {u.get('userid', ''): u for u in all_users}

        all_trucks = db.get_all(trucks.table_name, trucks.json_fields)
        trucks_map = {t.get('truckid', ''): t for t in all_trucks}

        all_driver_routes = db.get_all(driver_routes.table_name, driver_routes.json_fields)

        completed_routes = db.get_by_filter(
            route_table.table_name,
            [["status", "=", "completed"], ["approved", "=", 1]],
            route_table.json_fields,
            order=["-created_at"]
        )

        route_map = {route.get('route_id', ''): route for route in completed_routes}

        report_data = []
        for driver_route in all_driver_routes:
            truckid = driver_route.get('truckid', '')
            if not truckid:
                continue
            if truck_id and truckid != truck_id:
                continue

            route_id = driver_route.get('route_id', '')
            route = route_map.get(route_id, {})
            if not route:
                continue

            completed_at = route.get('completed_at')
            if completed_at:
                if from_date_obj and completed_at < from_date_obj:
                    continue
                if to_date_obj and completed_at > to_date_obj:
                    continue

            truck = trucks_map.get(truckid, {})
            truck_number = truck.get('registration_number') or truck.get('truck_number') or truckid or 'N/A'

            driverids = driver_route.get('driverid', [])
            if not isinstance(driverids, list):
                driverids = [driverids] if driverids else []

            driver1 = get_driver_name(driverids[0], users_map) if len(driverids) > 0 else 'N/A'
            driver2 = get_driver_name(driverids[1], users_map) if len(driverids) > 1 else 'N/A'

            route_distance = truck.get('distance', 0)
            actual_distance = truck.get('actual_distance', route_distance)
            extra_distance = calculate_distance_difference(route_distance, actual_distance)

            damage_reported = truck.get('damage_reported', False)
            damage_status = "Yes" if damage_reported else "No"
            truck_status = truck.get('status', 'Active')

            report_data.append({
                'truck_number': truck_number,
                'date': completed_at.strftime('%d/%m/%Y') if completed_at else 'N/A',
                'route_assigned': route.get('route_name', 'N/A'),
                'driver_1': driver1,
                'driver_2': driver2,
                'total_distance_km': round(float(route_distance) if route_distance else 0, 2),
                'actual_distance_km': round(float(actual_distance) if actual_distance else 0, 2),
                'extra_distance_km': extra_distance,
                'damage_reported': damage_status,
                'status': truck_status,
                'truckid': truckid,
                'route_id': route_id
            })

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Truck Report"

        # Add header with report info
        ws['A1'] = "Truck Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:K1')

        # Add date range
        ws['A2'] = f"Report Date Range: {from_date} to {to_date}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:K2')

        # Add summary statistics
        total_trucks = len(set([item.get('truckid', '') for item in report_data if item.get('truckid')]))
        total_trips = len(report_data)
        total_distance = sum([item.get('actual_distance_km', 0) for item in report_data])

        ws['A3'] = "Summary Statistics"
        ws['A3'].font = Font(size=12, bold=True)

        ws['A4'] = "Total Trucks:"
        ws['B4'] = total_trucks
        ws['A5'] = "Total Trips:"
        ws['B5'] = total_trips
        ws['A6'] = "Total Distance (KM):"
        ws['B6'] = round(total_distance, 2)

        # Add table header
        headers = ['Truck Number', 'Date', 'Route Assigned', 'Driver 1', 'Driver 2',
                   'Total Distance (KM)', 'Actual Distance (KM)', 'Extra Distance (KM)',
                   'Damage Reported', 'Status']
        header_fill = PatternFill(start_color="FF9500", end_color="FF9500", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add data rows
        for row_num, item in enumerate(report_data, 9):
            ws.cell(row=row_num, column=1).value = item.get('truck_number', 'N/A')
            ws.cell(row=row_num, column=2).value = item.get('date', 'N/A')
            ws.cell(row=row_num, column=3).value = item.get('route_assigned', 'N/A')
            ws.cell(row=row_num, column=4).value = item.get('driver_1', 'N/A')
            ws.cell(row=row_num, column=5).value = item.get('driver_2', 'N/A')
            ws.cell(row=row_num, column=6).value = item.get('total_distance_km', 0)
            ws.cell(row=row_num, column=7).value = item.get('actual_distance_km', 0)
            ws.cell(row=row_num, column=8).value = item.get('extra_distance_km', 0)
            ws.cell(row=row_num, column=9).value = item.get('damage_reported', 'No')
            ws.cell(row=row_num, column=10).value = item.get('status', 'Active')

            for col_num in range(1, 11):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 12

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'truck_report_{from_date}_to_{to_date}.xlsx'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

@bp_truck_report.route('/truck-report/export/pdf', methods=['GET'])
@jwt_required
def export_truck_report_pdf():
    """
    Export truck report to PDF format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - truck_id: specific truck ID (optional)
    - company_logo: URL to company logo (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export truck reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        truck_id = request.args.get('truck_id', '')
        company_logo = request.args.get('company_logo', '')

        # Build report data
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date else None
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date else None
        if to_date_obj:
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)

        all_users = db.get_all(users.table_name, users.json_fields)
        users_map = {u.get('userid', ''): u for u in all_users}

        all_trucks = db.get_all(trucks.table_name, trucks.json_fields)
        trucks_map = {t.get('truckid', ''): t for t in all_trucks}

        all_driver_routes = db.get_all(driver_routes.table_name, driver_routes.json_fields)

        completed_routes = db.get_by_filter(
            route_table.table_name,
            [["status", "=", "completed"], ["approved", "=", 1]],
            route_table.json_fields,
            order=["-created_at"]
        )

        route_map = {route.get('route_id', ''): route for route in completed_routes}

        report_data = []
        for driver_route in all_driver_routes:
            truckid = driver_route.get('truckid', '')
            if not truckid:
                continue
            if truck_id and truckid != truck_id:
                continue

            route_id = driver_route.get('route_id', '')
            route = route_map.get(route_id, {})
            if not route:
                continue

            completed_at = route.get('completed_at')
            if completed_at:
                if from_date_obj and completed_at < from_date_obj:
                    continue
                if to_date_obj and completed_at > to_date_obj:
                    continue

            truck = trucks_map.get(truckid, {})
            truck_number = truck.get('registration_number') or truck.get('truck_number') or truckid or 'N/A'

            driverids = driver_route.get('driverid', [])
            if not isinstance(driverids, list):
                driverids = [driverids] if driverids else []

            driver1 = get_driver_name(driverids[0], users_map) if len(driverids) > 0 else 'N/A'
            driver2 = get_driver_name(driverids[1], users_map) if len(driverids) > 1 else 'N/A'

            route_distance = truck.get('distance', 0)
            actual_distance = truck.get('actual_distance', route_distance)
            extra_distance = calculate_distance_difference(route_distance, actual_distance)

            damage_reported = truck.get('damage_reported', False)
            damage_status = "Yes" if damage_reported else "No"
            truck_status = truck.get('status', 'Active')

            report_data.append({
                'truck_number': truck_number,
                'date': completed_at.strftime('%d/%m/%Y') if completed_at else 'N/A',
                'route_assigned': route.get('route_name', 'N/A'),
                'driver_1': driver1,
                'driver_2': driver2,
                'total_distance_km': round(float(route_distance) if route_distance else 0, 2),
                'actual_distance_km': round(float(actual_distance) if actual_distance else 0, 2),
                'extra_distance_km': extra_distance,
                'damage_reported': damage_status,
                'status': truck_status,
                'truckid': truckid,
                'route_id': route_id
            })

        # Create PDF
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch, pageCompression=1)
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#FF9500'),
            spaceAfter=10,
            alignment=1
        )

        # Add company logo if provided
        if company_logo:
            try:
                logo = Image(company_logo, width=1*inch, height=1*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass

        # Add title
        elements.append(Paragraph("Truck Report", title_style))
        elements.append(Paragraph(f"<b>Report Date Range:</b> {from_date} to {to_date}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Add summary statistics
        total_trucks = len(set([item.get('truckid', '') for item in report_data if item.get('truckid')]))
        total_trips = len(report_data)
        total_distance = sum([item.get('actual_distance_km', 0) for item in report_data])

        summary_text = f"""
        <b>Summary Statistics</b><br/>
        Total Trucks: {total_trucks}<br/>
        Total Trips: {total_trips}<br/>
        Total Distance: {round(total_distance, 2)} KM
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Create report table with wrapped text
        normal_style = styles['Normal']
        normal_style.fontSize = 8
        normal_style.leading = 10

        table_data = [[Paragraph(text, normal_style) for text in ['Truck Number', 'Date', 'Route Assigned', 'Driver 1', 'Driver 2',
                      'Total Distance', 'Actual Distance', 'Extra Distance', 'Damage', 'Status']]]

        for item in report_data:
            table_data.append([
                Paragraph(str(item.get('truck_number', 'N/A')), normal_style),
                Paragraph(str(item.get('date', 'N/A')), normal_style),
                Paragraph(str(item.get('route_assigned', 'N/A')), normal_style),
                Paragraph(str(item.get('driver_1', 'N/A')), normal_style),
                Paragraph(str(item.get('driver_2', 'N/A')), normal_style),
                Paragraph(f"{item.get('total_distance_km', 0)} KM", normal_style),
                Paragraph(f"{item.get('actual_distance_km', 0)} KM", normal_style),
                Paragraph(f"{item.get('extra_distance_km', 0)} KM", normal_style),
                Paragraph(str(item.get('damage_reported', 'No')), normal_style),
                Paragraph(str(item.get('status', 'Active')), normal_style)
            ])

        # Create table
        table = Table(table_data, colWidths=[1.0*inch, 0.85*inch, 1.1*inch, 0.95*inch,
                                             0.95*inch, 1.0*inch, 1.0*inch, 1.0*inch,
                                             0.85*inch, 0.85*inch])

        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)

        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'truck_report_{from_date}_to_{to_date}.pdf'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)
