# Admin Driver Report Module

from flask import Blueprint, request, jsonify, make_response
from utils.exceptionlogging import ExceptionLogging
import traceback
from utils.jwt import jwt_required
from datetime import datetime, timedelta
import pytz
from utils import db
from utils.schemas import users, routes as route_table, driver_routes, log_table, trucks
from utils.globalconstants import CustomException
from utils.logging import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib import colors
from io import BytesIO
import json
from flask import send_file

bp_admin_report = Blueprint('bp_admin_report', __name__)

def get_driver_name(driverid, users_map):
    """Helper to get driver name from users map"""
    user = users_map.get(driverid, {})
    return user.get('name', 'Unknown')

def get_truck_name(truckid, trucks_map):
    """Helper to get truck registration number from trucks map"""
    truck = trucks_map.get(truckid, {})
    return truck.get('registration_number') or truckid or 'N/A'

def calculate_total_time(start_time, end_time):
    """Calculate total time taken in hours and minutes"""
    if not start_time or not end_time:
        return "N/A"
    try:
        diff = end_time - start_time
        hours = diff.seconds // 3600
        minutes = (diff.seconds % 3600) // 60
        return f"{hours} Hours {minutes} Minutes"
    except:
        return "N/A"

def parse_datetime(date_str):
    """Parse datetime string"""
    try:
        return datetime.fromisoformat(date_str)
    except:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except:
            return None

def build_driver_report_data(from_date_str, to_date_str, driver_id):
    """Helper function to build driver report data"""
    # Parse dates
    from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if from_date_str else None
    to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=pytz.UTC) if to_date_str else None

    if to_date:
        to_date = to_date.replace(hour=23, minute=59, second=59)

    # Get all users and trucks for mapping
    all_users = db.get_all(users.table_name, users.json_fields)
    all_trucks = db.get_all(trucks.table_name, trucks.json_fields)

    users_map = {u.get('userid', ''): u for u in all_users}
    trucks_map = {t.get('truckid', ''): t for t in all_trucks}

    # Get completed routes with driver assignments
    completed_routes = db.get_by_filter(
        route_table.table_name,
        [["status", "=", "completed"], ["approved", "=", 1]],
        route_table.json_fields,
        order=["-created_at"]
    )

    # Get all driver routes
    all_driver_routes = db.get_all(
        driver_routes.table_name,
        driver_routes.json_fields
    )

    # Build route map
    route_map = {dr.get('route_id', ''): dr for dr in all_driver_routes}

    # Build report data and calculate statistics
    report_data = []
    total_routes = 0
    total_diversions = 0
    drivers_set = set()

    for route in completed_routes:
        route_id = route.get('route_id', '')
        driver_route = route_map.get(route_id, {})

        if not driver_route:
            continue

        driverids = driver_route.get('driverid', [])
        if not isinstance(driverids, list):
            driverids = [driverids]

        # Filter by date range
        completed_at = route.get('completed_at')
        if completed_at:
            if from_date and completed_at < from_date:
                continue
            if to_date and completed_at > to_date:
                continue

        # Update statistics (before driver filter to get accurate totals)
        total_routes += 1
        drivers_set.update(driverids)
        total_diversions += route.get('diversions', 0)

        # Filter by specific driver if provided (for report data only)
        if driver_id and driver_id not in driverids:
            continue

        # Get driver names
        driver_names = [get_driver_name(did, users_map) for did in driverids]

        # Get truck info
        truckid = driver_route.get('truckid', '')
        truck_name = get_truck_name(truckid, trucks_map)

        # Get route start and end times
        start_time = route.get('created_at')
        end_time = route.get('completed_at')

        # Calculate total time
        total_time = calculate_total_time(start_time, end_time)

        # Get checkpoints covered
        checkpoints_covered = driver_route.get('checkpoints_covered', [])
        checkpoint_count = len(checkpoints_covered)

        # Get total checkpoints from route
        route_checkpoints = route.get('checkpoints', [])
        total_checkpoints = len(route_checkpoints)
        checkpoints_info = f"{checkpoint_count}/{total_checkpoints}"

        # Calculate completion percentage
        completion_percentage = 0
        if total_checkpoints > 0:
            completion_percentage = round((checkpoint_count / total_checkpoints) * 100)

        # Determine diversion status
        has_diversion = "Yes" if route.get('diversions', 0) > 0 else "No"

        report_data.append({
            'driver_name': ', '.join(driver_names),
            'assigned_route': route.get('route_name', 'N/A'),
            'assigned_truck': truck_name,
            'route_start_time': start_time.strftime('%d-%b-%Y %I:%M %p') if start_time else 'N/A',
            'route_end_time': end_time.strftime('%d-%b-%Y %I:%M %p') if end_time else 'N/A',
            'total_time_taken': total_time,
            'checkpoints_covered': checkpoints_info,
            'completion_percentage': f"{completion_percentage}%",
            'diversion': has_diversion,
            'diversion_count': route.get('diversions', 0),
            'status': driver_route.get('status', 'Unknown').capitalize(),
            'route_id': route_id,
            'driver_ids': driverids
        })

    return {
        'report_data': report_data,
        'summary': {
            'total_drivers': len(drivers_set),
            'total_routes_completed': total_routes,
            'total_diversions': total_diversions
        }
    }

@bp_admin_report.route('/driver-report/summary', methods=['GET'])
@jwt_required
def get_driver_report_summary():
    """
    Get driver report summary statistics
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access driver reports.")

        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')

        result_data = build_driver_report_data(from_date_str, to_date_str, '')

        payload.update({
            "message": "Driver report summary fetched successfully.",
            "summary": result_data['summary'],
            "from_date": from_date_str,
            "to_date": to_date_str
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

@bp_admin_report.route('/driver-report', methods=['GET'])
@jwt_required
def get_driver_report():
    """
    Get detailed driver report with report data
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - driver_id: specific driver ID (optional, if not provided fetches all drivers)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can access driver reports.")

        from_date_str = request.args.get('from_date', '')
        to_date_str = request.args.get('to_date', '')
        driver_id = request.args.get('driver_id', '')

        result_data = build_driver_report_data(from_date_str, to_date_str, driver_id)

        payload.update({
            "message": "Driver report fetched successfully.",
            "report_data": result_data['report_data'],
            "total_records": len(result_data['report_data']),
            "from_date": from_date_str,
            "to_date": to_date_str,
            "driver_id_filter": driver_id
        })
        result = 200

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

    return make_response(jsonify(payload), result)

@bp_admin_report.route('/driver-report/export/excel', methods=['GET'])
@jwt_required
def export_report_excel():
    """
    Export driver report to Excel format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - driver_id: specific driver ID (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export driver reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        driver_id = request.args.get('driver_id', '')

        result_data = build_driver_report_data(from_date, to_date, driver_id)
        report_data = result_data['report_data']

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Driver Report"

        # Add header with report info
        ws['A1'] = "Driver Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')

        # Add date range
        ws['A2'] = f"Report Date Range: {from_date} to {to_date}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:J2')

        # Add summary statistics
        total_drivers = len(set([item.get('driver_ids', [])[0] for item in report_data if item.get('driver_ids')]))
        total_routes = len(report_data)
        total_diversions = sum([item.get('diversion_count', 0) for item in report_data])

        ws['A3'] = "Summary Statistics"
        ws['A3'].font = Font(size=12, bold=True)

        ws['A4'] = "Total Drivers:"
        ws['B4'] = total_drivers
        ws['A5'] = "Total Routes Completed:"
        ws['B5'] = total_routes
        ws['A6'] = "Total Diversions:"
        ws['B6'] = total_diversions

        # Add table header
        headers = ['Driver Name', 'Assigned Route', 'Assigned Truck', 'Route Start Time',
                   'Route End Time', 'Total Time Taken', 'Checkpoints Covered',
                   'Completion %', 'Diversion', 'Status']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add data rows
        for row_num, item in enumerate(report_data, 9):
            ws.cell(row=row_num, column=1).value = item.get('driver_name', 'N/A')
            ws.cell(row=row_num, column=2).value = item.get('assigned_route', 'N/A')
            ws.cell(row=row_num, column=3).value = item.get('assigned_truck', 'N/A')
            ws.cell(row=row_num, column=4).value = item.get('route_start_time', 'N/A')
            ws.cell(row=row_num, column=5).value = item.get('route_end_time', 'N/A')
            ws.cell(row=row_num, column=6).value = item.get('total_time_taken', 'N/A')
            ws.cell(row=row_num, column=7).value = item.get('checkpoints_covered', 'N/A')
            ws.cell(row=row_num, column=8).value = item.get('completion_percentage', 'N/A')
            ws.cell(row=row_num, column=9).value = item.get('diversion', 'No')
            ws.cell(row=row_num, column=10).value = item.get('status', 'Completed')

            for col_num in range(1, 11):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'driver_report_{from_date}_to_{to_date}.xlsx'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)

@bp_admin_report.route('/driver-report/export/pdf', methods=['GET'])
@jwt_required
def export_report_pdf():
    """
    Export driver report to PDF format
    Query Parameters:
    - from_date: YYYY-MM-DD (start date)
    - to_date: YYYY-MM-DD (end date)
    - driver_id: specific driver ID (optional)
    - company_logo: URL to company logo (optional)
    """
    payload, result = {
        "message": "Oops! Something went wrong. Please try again."
    }, 400
    try:
        user = request.user
        if user['role'] != 'admin':
            raise CustomException("Only admin can export driver reports.")

        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        driver_id = request.args.get('driver_id', '')
        company_logo = request.args.get('company_logo', '')

        result_data = build_driver_report_data(from_date, to_date, driver_id)
        report_data = result_data['report_data']

        # Create PDF
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch, pageCompression=1)
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=10,
            alignment=1
        )

        # Add company logo if provided
        if company_logo:
            try:
                logo = Image(company_logo, width=1*inch, height=1*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass

        # Add title
        elements.append(Paragraph("Driver Report", title_style))
        elements.append(Paragraph(f"<b>Report Date Range:</b> {from_date} to {to_date}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Add summary statistics
        total_drivers = len(set([item.get('driver_ids', [])[0] for item in report_data if item.get('driver_ids')]))
        total_routes = len(report_data)
        total_diversions = sum([item.get('diversion_count', 0) for item in report_data])

        summary_text = f"""
        <b>Summary Statistics</b><br/>
        Total Drivers: {total_drivers}<br/>
        Total Routes Completed: {total_routes}<br/>
        Total Diversions: {total_diversions}
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Create report table with wrapped text
        normal_style = styles['Normal']
        normal_style.fontSize = 8
        normal_style.leading = 10

        table_data = [[Paragraph(text, normal_style) for text in ['Driver Name', 'Assigned Route', 'Assigned Truck', 'Route Start Time',
                      'Route End Time', 'Total Time Taken', 'Checkpoints Covered',
                      'Completion %', 'Diversion', 'Status']]]

        for item in report_data:
            table_data.append([
                Paragraph(str(item.get('driver_name', 'N/A')), normal_style),
                Paragraph(str(item.get('assigned_route', 'N/A')), normal_style),
                Paragraph(str(item.get('assigned_truck', 'N/A')), normal_style),
                Paragraph(str(item.get('route_start_time', 'N/A')), normal_style),
                Paragraph(str(item.get('route_end_time', 'N/A')), normal_style),
                Paragraph(str(item.get('total_time_taken', 'N/A')), normal_style),
                Paragraph(str(item.get('checkpoints_covered', 'N/A')), normal_style),
                Paragraph(str(item.get('completion_percentage', 'N/A')), normal_style),
                Paragraph(str(item.get('diversion', 'No')), normal_style),
                Paragraph(str(item.get('status', 'Completed')), normal_style)
            ])

        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 0.9*inch, 1.1*inch,
                                             1.1*inch, 1.1*inch, 0.95*inch, 0.9*inch,
                                             0.8*inch, 0.8*inch])

        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)

        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'driver_report_{from_date}_to_{to_date}.pdf'
        )

    except CustomException as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        payload.update({"message": str(e.message)})
        return make_response(jsonify(payload), result)
    except Exception as e:
        ExceptionLogging.LogException(traceback.format_exc(), e)
        return make_response(jsonify(payload), result)
